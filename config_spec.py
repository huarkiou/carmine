"""Collect autohome parameter config tables for all series, organized by manufacturer/brand/model."""
import requests
import time
import re
import os
import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from collect import (
    BRAND_TO_MANUFACTURER, clean_manu_name, HEADERS, fetch_brand_map,
    fill_missing_brands, create_manu_map, CATEGORIES, MONTHS,
)

API_RANK = "https://www.autohome.com.cn/web-main/car/rank/getList"
API_CONFIG = "https://www.autohome.com.cn/web-main/car/param/getParamConf"
OUTPUT_DIR = "D:/Users/huarkiou/Downloads/配置表"

ONLY_ON_SALE = True  # True=仅在售年款(lazyload:0), False=全部年款


def collect_all_series():
    """Collect all unique series from ranking data."""
    all_series = {}  # seriesid -> {name, brandid}
    for cat_big, subcats in CATEGORIES.items():
        for sub_name, levelid in subcats:
            params = {
                "from": 28, "pm": 2, "pluginversion": "11.75.8",
                "model": 1, "channel": 0, "pageindex": 1, "pagesize": 50,
                "typeid": 1, "subranktypeid": 1, "levelid": levelid,
                "price": "0-9000", "date": "2026-04",
            }
            try:
                r = requests.get(API_RANK, params=params, headers=HEADERS, timeout=15)
                r.encoding = "utf-8"
                for item in r.json().get("result", {}).get("list", []):
                    sid = str(item.get("seriesid", ""))
                    if sid and sid not in all_series:
                        all_series[sid] = {
                            "name": item.get("seriesname", ""),
                            "brandid": item.get("brandid", 0),
                        }
            except Exception:
                pass
            time.sleep(0.2)
    print(f"Collected {len(all_series)} unique series")
    return all_series


def get_config(seriesid):
    """Fetch parameter config for a series (all years, no filter)."""
    try:
        url = f"{API_CONFIG}?mode=1&site=1&seriesid={seriesid}"
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.encoding = "utf-8"
        return r.json().get("result", {})
    except Exception:
        return {}


def parse_config(result):
    """Parse config API result into structured format.
    
    Returns: {year_name: [(group_name, [(param_name, [spec_values...]), ...]), ...]}
    """
    titlelist = result.get("titlelist", [])
    datalist = result.get("datalist", [])
    conditionlist = result.get("conditionlist", [])

    # Build flat param index: titleid -> (group_name, item_name)
    param_index = {}
    for group in titlelist:
        group_name = group.get("itemtype", "")
        for item in group.get("items", []):
            tid = item.get("titleid")
            if tid is not None:
                param_index[tid] = (group_name, item.get("itemname", ""))

    # Identify year options
    year_options = {}
    for cond in conditionlist:
        if cond.get("typevalue") == "year":
            for y in cond.get("list", []):
                if not ONLY_ON_SALE or y.get("lazyload") == 0:
                    year_options[y.get("id")] = y.get("name", "")

    # Group specs by year (year is last element in condition array)
    year_specs = defaultdict(list)
    for spec in datalist:
        cond = spec.get("condition", [])
        spec_year = cond[-1] if isinstance(cond, list) and cond else ""
        if not spec_year or not spec_year.isdigit():
            continue
        if spec_year not in year_options:
            continue
        year_specs[spec_year].append(spec)

    # Build output: {year_name: [(group_name, [(param_name, [values]), ...]), ...]}
    output = {}
    for year_id, specs in year_specs.items():
        year_name = year_options.get(year_id, f"{year_id}款")
        # Build param row data
        param_rows = []  # [(group_name, param_name, [val1, val2, ...])]
        last_group = None
        for group in titlelist:
            group_name = group.get("itemtype", "")
            for item in group.get("items", []):
                tid = item.get("titleid")
                pname = item.get("itemname", "")
                values = []
                for spec in specs:
                    pcl = spec.get("paramconflist", [])
                    # Find matching param value by titleid
                    val = ""
                    for p in pcl:
                        if p.get("titleid") == tid:
                            val = p.get("itemname", "")
                            break
                    values.append(val)
                param_rows.append((group_name, pname, values))

        # Build spec names
        spec_names = [s.get("specname", "") for s in specs]
        output[year_name] = (spec_names, param_rows)

    return output, year_options


def write_config_xlsx(filepath, config_data):
    """Write config data to xlsx. Uses temp file + atomic rename to prevent partial writes."""
    tmp_path = filepath + ".tmp"
    # Clean up any stale temp file from previous interrupted run
    if os.path.exists(tmp_path):
        os.remove(tmp_path)
    wb = Workbook()
    wb.remove(wb.active)

    hf = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=9)
    calign = Alignment(vertical="center", wrap_text=True)
    group_fill = PatternFill("solid", fgColor="D9E2F3")
    group_font = Font(name="Arial", bold=True, size=9)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for year_name, (spec_names, param_rows) in config_data.items():
        # Sheet name: max 31 chars
        sheet_name = year_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        if not spec_names or not param_rows:
            ws.cell(row=1, column=1, value="暂无数据").font = cf
            continue

        num_specs = len(spec_names)
        total_cols = 1 + num_specs

        # Write spec name header row
        ws.cell(row=1, column=1, value="参数").font = hf
        ws.cell(row=1, column=1).fill = hfill
        ws.cell(row=1, column=1).alignment = halign
        ws.cell(row=1, column=1).border = border
        for ci, sname in enumerate(spec_names):
            cell = ws.cell(row=1, column=2 + ci, value=sname)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = halign
            cell.border = border

        # Write param rows
        row_idx = 2
        for group_name, pname, values in param_rows:
            # Check if all values are empty (non-applicable param)
            cell = ws.cell(row=row_idx, column=1, value=pname)
            cell.font = group_font
            cell.alignment = calign
            cell.border = border
            for ci, val in enumerate(values):
                cell = ws.cell(row=row_idx, column=2 + ci, value=val)
                cell.font = cf
                cell.alignment = calign
                cell.border = border
            row_idx += 1

        # Set column widths
        ws.column_dimensions['A'].width = 22
        for ci in range(num_specs):
            col_letter = get_column_letter(2 + ci)
            ws.column_dimensions[col_letter].width = 18

        ws.freeze_panes = "B2"

    wb.save(tmp_path)
    # Atomic rename: only replace final file after successful write
    if os.path.exists(filepath):
        os.remove(filepath)
    os.rename(tmp_path, filepath)
    return True


def main():
    print("=== Step 1: Fetching brand list ===")
    brand_map = fetch_brand_map()
    manu_map = create_manu_map(brand_map)

    print("=== Step 2: Collecting all series ===")
    all_series = collect_all_series()

    # Build brandid->seriesid mapping for missing brands
    series_by_brand = defaultdict(list)
    for sid, info in all_series.items():
        bid = info["brandid"]
        if bid:
            series_by_brand[bid].append(sid)

    print("=== Step 3: Filling missing brand names ===")
    fill_missing_brands(brand_map, manu_map, series_by_brand)
    # Re-apply explicit manufacturer mapping
    new_manu = create_manu_map(brand_map)
    previously_missing = {bid for bid in series_by_brand if bid not in brand_map}
    for bid in previously_missing:
        if bid in new_manu:
            manu_map[bid] = new_manu[bid]

    print(f"\n=== Step 4: Fetching config for {len(all_series)} series ===")
    stats = {"success": 0, "empty": 0, "error": 0, "skipped": 0}

    for i, (sid, info) in enumerate(sorted(all_series.items())):
        series_name = info["name"]
        brandid = info["brandid"]
        brand_name = brand_map.get(brandid, f"品牌{brandid}")
        manufacturer = manu_map.get(brandid, clean_manu_name(brand_name))

        # Build output path and skip if already exists
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', series_name)
        dir_path = os.path.join(OUTPUT_DIR, manufacturer, brand_name)
        filepath = os.path.join(dir_path, f"{safe_name}.xlsx")
        if os.path.exists(filepath):
            # Check for stale temp file from interrupted previous run
            tmp_path = filepath + ".tmp"
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
                os.remove(filepath)  # Re-process: existing file may be incomplete
            else:
                stats["skipped"] += 1
                if stats["skipped"] % 50 == 0:
                    print(f"  skipped {stats['skipped']} existing files...")
                continue

        print(f"  [{i+1}/{len(all_series)}] {series_name} (sid={sid})", end=" ")

        result = get_config(sid)
        if not result:
            print("ERROR")
            stats["error"] += 1
            time.sleep(0.3)
            continue

        config_data, year_options = parse_config(result)
        if not config_data:
            print("no on-sale data" if ONLY_ON_SALE else "empty")
            stats["empty"] += 1
            time.sleep(0.3)
            continue

        # Build output path
        dir_path = os.path.join(OUTPUT_DIR, manufacturer, brand_name)
        os.makedirs(dir_path, exist_ok=True)
        # Sanitize filename
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', series_name)
        filepath = os.path.join(dir_path, f"{safe_name}.xlsx")

        ok = write_config_xlsx(filepath, config_data)
        if ok:
            years = list(config_data.keys())
            print(f"-> {len(years)} years: {', '.join(years)} | {manufacturer}/{brand_name}/{safe_name}.xlsx")
            stats["success"] += 1
        else:
            print("WRITE ERROR")
            stats["error"] += 1

        time.sleep(0.15)

    print(f"\nDone: {stats['success']} success, {stats['empty']} empty, {stats['error']} errors, {stats['skipped']} skipped")
    print(f"Output: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
