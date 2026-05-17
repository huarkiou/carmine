"""Verify output xlsx files for data quality issues.

Usage:
    uv run python tools/verify.py <file.xlsx>   # auto-detect type
    uv run python tools/verify.py                # default sales file
"""
import sys
from openpyxl import load_workbook


def verify_sales(filepath):
    """Check sales ranking xlsx for brand placeholder & manufacturer suffix issues."""
    wb = load_workbook(filepath)
    issues = []
    stats = {"total": 0, "brands": set(), "manus": set()}

    for name in wb.sheetnames:
        ws = wb[name]
        for row in range(2, ws.max_row + 1):
            model = str(ws.cell(row=row, column=2).value or '')
            brand = str(ws.cell(row=row, column=3).value or '')
            manu = str(ws.cell(row=row, column=4).value or '')

            stats["total"] += 1
            stats["brands"].add(brand)
            stats["manus"].add(manu)

            if '品牌' in brand:
                issues.append(f"[BRAND] {model}: brand={brand}")
            if '品牌' in manu:
                issues.append(f"[MANU] {model}: manu={manu}")
            for suffix in ['汽车', '集团']:
                if manu.endswith(suffix) and manu not in ['上汽通用五菱', '东风日产', '鸿蒙智行']:
                    issues.append(f"[SUFFIX] {model}: manu={manu} ends with '{suffix}'")

    _print_results(stats, issues)


def verify_config(filepath):
    """Check config xlsx: count empty params per sheet."""
    wb = load_workbook(filepath)
    stats = {"sheets": len(wb.sheetnames)}

    for name in wb.sheetnames:
        ws = wb[name]
        empty_rows = 0
        for row in range(2, ws.max_row + 1):
            all_empty = True
            for col in range(2, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None and str(v).strip() not in ('', '-'):
                    all_empty = False
                    break
            if all_empty:
                empty_rows += 1
        print(f"  Sheet '{name}': {ws.max_row-1} params x {ws.max_column-1} specs, {empty_rows} all-empty rows")


def _print_results(stats, issues):
    print(f"Rows: {stats.get('total', 0)}")
    if 'brands' in stats:
        print(f"Unique brands: {len(stats['brands'])}")
        print(f"Unique manufacturers: {len(stats['manus'])}")
    print(f"Issues found: {len(issues)}")
    for i in issues[:20]:
        print(f"  {i}")
    if len(issues) > 20:
        print(f"  ... and {len(issues)-20} more")


def detect_type(filepath):
    """Detect if file is sales or config xlsx based on headers."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    h1 = str(ws.cell(row=1, column=1).value or '')
    h2 = str(ws.cell(row=1, column=2).value or '')
    if h1 == '排名' and '车型' in h2:
        return 'sales'
    if h1 == '参数':
        return 'config'
    return 'sales'  # default


if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else 'output/latest/汽车销量排行-近6个月.xlsx'
    ftype = detect_type(path)
    print(f"File: {path} (detected: {ftype})")

    if ftype == 'config':
        verify_config(path)
    else:
        verify_sales(path)
