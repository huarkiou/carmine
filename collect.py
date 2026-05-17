"""Collect autohome 6-month aggregated sales data by vehicle category."""
import requests
import time
import re
from collections import defaultdict
from urllib.parse import unquote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API = "https://www.autohome.com.cn/web-main/car/rank/getList"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://www.autohome.com.cn/rank/1-1-0-0_9000-x-x-x/2026-04.html",
}
OUTPUT = "D:/Users/huarkiou/Downloads/汽车销量排行-近6个月.xlsx"

MONTHS = ["2025-11", "2025-12", "2026-01", "2026-02", "2026-03", "2026-04"]

CATEGORIES = {
    "轿车": [
        ("微型车", "1"), ("小型车", "2"), ("紧凑型车", "3"),
        ("中型车", "4"), ("中大型车", "5"), ("大型车", "6"),
    ],
    "SUV": [
        ("小型SUV", "16"), ("紧凑型SUV", "17"), ("中型SUV", "18"),
        ("中大型SUV", "19"), ("大型SUV", "20"),
    ],
    "MPV": [
        ("紧凑型MPV", "21"), ("中型MPV", "22"),
        ("中大型MPV", "23"), ("大型MPV", "24"),
    ],
}

# brand_name -> manufacturer_name mapping for key Chinese auto groups
BRAND_TO_MANUFACTURER = {
    # 比亚迪集团
    "比亚迪": "比亚迪",
    "方程豹": "比亚迪",
    "腾势": "比亚迪",
    "仰望": "比亚迪",
    # 吉利集团
    "吉利汽车": "吉利",
    "吉利银河": "吉利",
    "吉利几何": "吉利",
    "领克": "吉利",
    "极氪": "吉利",
    "睿蓝汽车": "吉利",
    # 长城集团
    "哈弗": "长城",
    "魏牌": "长城",
    "坦克": "长城",
    "欧拉": "长城",
    # 长安集团
    "长安": "长安",
    "长安启源": "长安",
    "深蓝汽车": "长安",
    "阿维塔": "长安",
    "长安欧尚": "长安",
    "长安凯程": "长安",
    # 奇瑞集团
    "奇瑞": "奇瑞",
    "奇瑞QQ": "奇瑞",
    "奇瑞风云": "奇瑞",
    "捷途": "奇瑞",
    "捷途山海": "奇瑞",
    "星途": "奇瑞",
    "iCAR": "奇瑞",
    "凯翼": "奇瑞",
    # 蔚来集团
    "蔚来": "蔚来",
    "乐道": "蔚来",
    "firefly萤火虫": "蔚来",
    # 小鹏/理想/零跑/小米
    "小鹏": "小鹏",
    "理想汽车": "理想",
    "零跑汽车": "零跑",
    "小米汽车": "小米",
    "特斯拉": "特斯拉",
    # 东风集团
    "东风风神": "东风",
    "东风奕派": "东风",
    "东风风行": "东风",
    "东风风度": "东风",
    "东风风光": "东风",
    "东风富康": "东风",
    "岚图汽车": "东风",
    "猛士": "东风",
    "蓝电": "东风",
    "纳米": "东风",
    # 上汽集团
    "荣威": "上汽",
    "名爵": "上汽",
    "智己汽车": "上汽",
    "大通": "上汽",
    # 上汽通用五菱
    "五菱汽车": "上汽通用五菱",
    "宝骏": "上汽通用五菱",
    # 广汽集团
    "广汽传祺": "广汽",
    "埃安": "广汽",
    "广汽昊铂": "广汽",
    # 北汽集团
    "ARCFOX极狐": "北汽",
    "北京汽车": "北汽",
    "北京汽车制造厂": "北汽",
    "北汽新能源": "北汽",
    "北京越野": "北汽",
    # 一汽集团
    "红旗": "一汽",
    "奔腾": "一汽",
    "捷达": "一汽",
    # 鸿蒙智行
    "鸿蒙智行": "鸿蒙智行",
    "AITO 问界": "赛力斯",
    "智界": "鸿蒙智行",
    "享界": "鸿蒙智行",
    "尊界": "鸿蒙智行",
    "尚界": "鸿蒙智行",
    # 江淮
    "江淮汽车": "江淮",
    "江淮瑞风": "江淮",
    "江淮钇为": "江淮",
    # 其他中国品牌
    "中国重汽VGV": "中国重汽",
    "SWM斯威汽车": "鑫源",
    "创维汽车": "创维",
    "曹操汽车": "曹操",
    "凌宝汽车": "凌宝",
    "鑫源汽车": "鑫源",
    "启辰": "东风日产",
    "212": "北汽",
    "ROX极石": "极石",
    # 合资/外资品牌
    "大众": "大众",
    "丰田": "丰田",
    "本田": "本田",
    "日产": "日产",
    "宝马": "宝马",
    "奔驰": "奔驰",
    "奥迪": "奥迪",
    "凯迪拉克": "凯迪拉克",
    "沃尔沃": "沃尔沃",
    "福特": "福特",
    "别克": "别克",
    "雪佛兰": "雪佛兰",
    "雪铁龙": "雪铁龙",
    "标致": "标致",
    "现代": "现代",
    "起亚": "起亚",
    "马自达": "马自达",
    "路虎": "路虎",
    "捷豹": "捷豹",
    "英菲尼迪": "英菲尼迪",
    "林肯": "林肯",
    "smart": "smart",
    "福田": "福田",
    "灵悉": "灵悉",
    "纵横": "纵横",
    "示界": "示界",
    "埃尚": "埃尚",
}


def fetch_brand_map():
    """Build brandid -> brandname mapping from brand monthly ranking."""
    brand_map = {}
    for page in range(1, 3):
        params = {
            "from": 28, "pm": 2, "pluginversion": "11.75.8",
            "model": 1, "channel": 0, "pageindex": page, "pagesize": 200,
            "typeid": 1, "subranktypeid": 3, "entitytype": "1071",
            "date": "2026-04",
        }
        try:
            r = requests.get(API, params=params, headers=HEADERS, timeout=15)
            r.encoding = "utf-8"
            data = r.json()
            items = data.get("result", {}).get("list", [])
            if not items:
                break
            for item in items:
                bname = item.get("seriesname", "")
                args = item.get("pvitem", {}).get("argvs", {})
                brandid = args.get("brandid")
                if not brandid:
                    m = re.search(r"brandid=(\d+)", item.get("linkurl", ""))
                    if m:
                        brandid = m.group(1)
                if brandid and bname and int(brandid) not in brand_map:
                    brand_map[int(brandid)] = bname
        except Exception as e:
            print(f"  Brand page {page} error: {e}")
        time.sleep(0.3)
    print(f"Fetched {len(brand_map)} brands")
    return brand_map


def clean_manu_name(name):
    """Strip common suffixes from manufacturer names."""
    if not name:
        return name
    for suffix in ["汽车制造厂", "汽车", "集团"]:
        if name.endswith(suffix) and len(name) > len(suffix):
            name = name[:-len(suffix)]
    # Handle specific: 东风日产 stays, 中国重汽 stays
    return name


def create_manu_map(brand_map):
    """Map brandid -> manufacturer_name using known brand-to-manufacturer relationships."""
    result = {}
    for brandid, brand_name in brand_map.items():
        manu = BRAND_TO_MANUFACTURER.get(brand_name)
        if not manu:
            for key in BRAND_TO_MANUFACTURER:
                if key in brand_name or brand_name in key:
                    manu = BRAND_TO_MANUFACTURER[key]
                    break
        if not manu:
            manu = clean_manu_name(brand_name)
        result[brandid] = manu
    return result


def fetch_series(levelid, month):
    """Fetch series ranking for one month and level."""
    params = {
        "from": 28, "pm": 2, "pluginversion": "11.75.8",
        "model": 1, "channel": 0, "pageindex": 1, "pagesize": 50,
        "typeid": 1, "subranktypeid": 1, "levelid": levelid,
        "price": "0-9000", "date": month,
    }
    try:
        r = requests.get(API, params=params, headers=HEADERS, timeout=15)
        r.encoding = "utf-8"
        data = r.json()
        return data.get("result", {}).get("list", [])
    except Exception as e:
        return []


def main():
    print("=== Step 1: Fetching brand list ===")
    brand_map = fetch_brand_map()

    print("=== Step 2: Building manufacturer mapping ===")
    manu_map = create_manu_map(brand_map)

    print("=== Step 3: Collecting 6-month sales data ===")
    all_data = {}

    for cat_big, subcats in CATEGORIES.items():
        print(f"\n--- {cat_big} ---")
        sub_data = defaultdict(lambda: defaultdict(lambda: {
            "name": "", "brandid": 0, "total_sales": 0,
            "price": "", "months": set(),
        }))

        for sub_name, levelid in subcats:
            for month in MONTHS:
                items = fetch_series(levelid, month)
                for item in items:
                    sid = str(item.get("seriesid", ""))
                    sales = item.get("salecount", 0) or 0
                    entry = sub_data[sub_name][sid]
                    entry["name"] = item.get("seriesname", "")
                    entry["brandid"] = item.get("brandid", 0)
                    entry["total_sales"] += int(sales)
                    entry["price"] = item.get("priceinfo", "")
                    entry["months"].add(month)
                time.sleep(0.25)
            print(f"  {sub_name}: {len(sub_data[sub_name])} series")

        all_data[cat_big] = sub_data

    print("\n=== Step 4: Aggregating and building output ===")
    output = {}
    unmapped = set()

    for cat_big, sub_data in all_data.items():
        output[cat_big] = []
        for sub_name, series_map in sub_data.items():
            rows = []
            for sid, info in series_map.items():
                brandid = info["brandid"]
                brand_name = brand_map.get(brandid)
                if not brand_name:
                    brand_name = f"品牌{brandid}"
                manufacturer = manu_map.get(brandid, brand_name)
                if manufacturer == brand_name and manufacturer not in BRAND_TO_MANUFACTURER:
                    unmapped.add(f"{brandid}:{brand_name}")
                rows.append({
                    "车型名称": info["name"],
                    "品牌": brand_name,
                    "主机厂": manufacturer,
                    "6个月总销量": info["total_sales"],
                    "价格区间": info["price"],
                    "子分类": sub_name,
                })
            rows.sort(key=lambda x: x["6个月总销量"], reverse=True)
            rows = rows[:50]
            for i, r in enumerate(rows):
                r["排名"] = i + 1
            output[cat_big].append((sub_name, rows))

    if unmapped:
        print(f"\nUnmapped brands (using brand name as manufacturer):")
        for u in sorted(unmapped):
            print(f"  {u}")

    print("\n=== Step 5: Writing Excel ===")
    write_excel(output)
    print(f"\nDone: {OUTPUT}")


def write_excel(output):
    wb = Workbook()
    wb.remove(wb.active)

    hf = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    calign = Alignment(vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    cols = ["排名", "车型名称", "品牌", "主机厂", "6个月总销量", "价格区间", "子分类"]
    widths = [8, 22, 16, 18, 14, 16, 14]

    for cat_big, cat_data in output.items():
        ws = wb.create_sheet(title=cat_big)

        for ci, cn in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=cn)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = halign
            cell.border = border
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        row_idx = 2
        for sub_name, rows in cat_data:
            for r in rows:
                for ci, cn in enumerate(cols, 1):
                    cell = ws.cell(row=row_idx, column=ci, value=r[cn])
                    cell.font = cf
                    cell.alignment = calign
                    cell.border = border
                row_idx += 1

        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{row_idx - 1}"
        ws.freeze_panes = "A2"

    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
