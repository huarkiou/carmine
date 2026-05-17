"""Excel output functions with shared styling constants."""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Shared styles
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Arial", size=10)
CELL_ALIGN = Alignment(vertical="center")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CONFIG_HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
CONFIG_CELL_FONT = Font(name="Arial", size=9)
CONFIG_CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
CONFIG_GROUP_FILL = PatternFill("solid", fgColor="D9E2F3")
CONFIG_GROUP_FONT = Font(name="Arial", bold=True, size=9)


def write_sales_excel(output, filepath):
    """Write aggregated sales ranking data to xlsx.
    
    output: {大类名: [(子分类名, [rows]), ...]}
    """
    wb = Workbook()
    wb.remove(wb.active)

    cols = ["排名", "车型名称", "品牌", "主机厂", "6个月总销量", "价格区间", "子分类"]
    widths = [8, 22, 16, 18, 14, 16, 14]

    for cat_big, cat_data in output.items():
        ws = wb.create_sheet(title=cat_big)

        for ci, cn in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=cn)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = BORDER
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        row_idx = 2
        for sub_name, rows in cat_data:
            for r in rows:
                for ci, cn in enumerate(cols, 1):
                    cell = ws.cell(row=row_idx, column=ci, value=r[cn])
                    cell.font = CELL_FONT
                    cell.alignment = CELL_ALIGN
                    cell.border = BORDER
                row_idx += 1

        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{row_idx - 1}"
        ws.freeze_panes = "A2"

    wb.save(filepath)


def write_config_xlsx(filepath, config_data):
    """Write parameter configuration table to xlsx with atomic temp-file write.
    
    config_data: {year_name: (spec_names, param_rows)}
    param_rows: [(group_name, param_name, [spec_values...]), ...]
    """
    tmp_path = filepath + ".tmp"
    if os.path.exists(tmp_path):
        os.remove(tmp_path)
    wb = Workbook()
    wb.remove(wb.active)

    for year_name, (spec_names, param_rows) in config_data.items():
        sheet_name = year_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        if not spec_names or not param_rows:
            ws.cell(row=1, column=1, value="暂无数据").font = CONFIG_CELL_FONT
            continue

        num_specs = len(spec_names)

        cell = ws.cell(row=1, column=1, value="参数")
        cell.font = CONFIG_HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        for ci, sname in enumerate(spec_names):
            cell = ws.cell(row=1, column=2 + ci, value=sname)
            cell.font = CONFIG_HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = BORDER

        row_idx = 2
        for group_name, pname, values in param_rows:
            cell = ws.cell(row=row_idx, column=1, value=pname)
            cell.font = CONFIG_GROUP_FONT
            cell.alignment = CONFIG_CELL_ALIGN
            cell.border = BORDER
            for ci, val in enumerate(values):
                cell = ws.cell(row=row_idx, column=2 + ci, value=val)
                cell.font = CONFIG_CELL_FONT
                cell.alignment = CONFIG_CELL_ALIGN
                cell.border = BORDER
            row_idx += 1

        ws.column_dimensions['A'].width = 22
        for ci in range(num_specs):
            ws.column_dimensions[get_column_letter(2 + ci)].width = 18
        ws.freeze_panes = "B2"

    wb.save(tmp_path)
    if os.path.exists(filepath):
        os.remove(filepath)
    os.rename(tmp_path, filepath)
    return True
